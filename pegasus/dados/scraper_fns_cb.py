import os
from os import path
import pathlib
from abc import ABC, abstractmethod
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import xlsxwriter
import openpyxl



# Class implementada pela Monique no repositório CoviDATA
class SeleniumDownloader(ABC):
    def __init__(self, diretorio_dados, url):
        self.driver = self.__configurar_chrome(diretorio_dados)
        self.driver.get(url)

    def download(self):
        self._executar()

        # Aguarda o download
        time.sleep(5)

        #self.driver.close()
        #self.driver.quit()

    @abstractmethod
    def _executar(self):
        pass

    def __configurar_chrome(self, diretorio_dados):
        if not path.exists(diretorio_dados):
            os.makedirs(diretorio_dados)

        chromeOptions = webdriver.ChromeOptions()
        prefs = {"download.default_directory": diretorio_dados}
        chromeOptions.add_experimental_option("prefs", prefs)
        #chromeOptions.add_argument('--headless')
        chromeOptions.add_argument('--start-maximized')


        driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chromeOptions)

        driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        params = {'cmd': 'Page.setDownloadBehavior',
                  'params': {'behavior': 'allow', 'downloadPath': diretorio_dados}}
        command_result = driver.execute("send_command", params)

        return driver


################################################################################

# Scraper FNS Contas Bancárias

################################################################################


diretorio_raiz = pathlib.Path().absolute()
diretorio_dados = diretorio_raiz.joinpath('dados')

estados_values_capitais = {'ACRE': '15',
                           'ALAGOAS': '46',
                           'AMAPA': '6',
                           'AMAZONAS': '37',
                           'BAHIA': '335',
                           'CEARA': '58',
                           'DISTRITO FEDERAL': '0',
                           'ESPIRITO SANTO': '77',
                           'GOIAS': '94',
                           'MARANHAO': '185',
                           'MATO GROSSO': '37',
                           'MATO GROSSO DO SUL': '19',
                           'MINAS GERAIS': '65',
                           'PARA': '18',
                           'PARAIBA': '93',
                           'PARANA': '94',
                           'PERNAMBUCO': '131',
                           'PIAUI': '215',
                           'RIO DE JANEIRO': '67',
                           'RIO GRANDE DO NORTE': '87',
                           'RIO GRANDE DO SUL': '325',
                           'RONDONIA': '36',
                           'RORAIMA': '2',
                           'SANTA CATARINA': '89',
                           'SAO PAULO': '564',
                           'SERGIPE': '2',
                           'TOCANTINS': '88'}

qtd_munic_estado = {'ACRE': 22,
                    'ALAGOAS': 102,
                    'AMAPA': 16,
                    'AMAZONAS': 62,
                    'BAHIA': 417,
                    'CEARA': 184,
                    'DISTRITO FEDERAL': 1,
                    'ESPIRITO SANTO': 78,
                    'GOIAS': 246,
                    'MARANHAO': 217,
                    'MATO GROSSO': 141,
                    'MATO GROSSO DO SUL': 79,
                    'MINAS GERAIS': 499,
                    'PARA': 144,
                    'PARAIBA': 223,
                    'PARANA': 399,
                    'PERNAMBUCO': 184,
                    'PIAUI': 224,
                    'RIO DE JANEIRO': 92,
                    'RIO GRANDE DO NORTE': 167,
                    'RIO GRANDE DO SUL': 497,
                    'RONDONIA': 52,
                    'RORAIMA': 15,
                    'SANTA CATARINA': 295,
                    'SAO PAULO': 645,
                    'SERGIPE': 75,
                    'TOCANTINS': 139}

# Define a classe referida como herdeira da class "SeleniumDownloader"
class FNS_CONTA_BANCARIA(SeleniumDownloader):

    # Sobrescreve o construtor da class "SeleniumDownloader"
    def __init__(self):
        super().__init__(path.join(diretorio_dados, 'fns'),
                         'https://consultafns.saude.gov.br/#/conta-bancaria')

    # Implementa localmente o método interno e vazio da class "SeleniumDownloader"
    def _executar(self):

        wait = WebDriverWait(self.driver, 20)

        time.sleep(5)

        try:

            workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

            workbook.save(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

        except:

            workbook = xlsxwriter.Workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

            workbook.close()

        workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

        sheets_names = workbook.sheetnames

        number_sheets = len(sheets_names)

        workbook.save(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

        for estado in list(estados_values_capitais.keys())[number_sheets - 1:]:

            df_estado = pd.DataFrame(columns=['UF', 'MUNICÍPIO', 'ESFERA', 'BANCO', 'AGÊNCIA', 'CONTA',
                                              'TIPO CONTA', 'CNPJ', 'ENTIDADE', 'VALOR SALDO'])

            time.sleep(5)

            select = Select(self.driver.find_element_by_id('estado'))
            select.select_by_visible_text(estado)

            qtd_estado = qtd_munic_estado[estado]

            for municipio in range(qtd_estado):

                time.sleep(1)

                municipio = str(municipio)

                select = Select(self.driver.find_element_by_id('municipio'))
                select.select_by_value(municipio)

                if municipio == estados_values_capitais[estado]:

                    for esfera in ['ESTADUAL', 'MUNICIPAL']:

                        time.sleep(1)

                        select = Select(self.driver.find_element_by_id('esfera'))
                        select.select_by_value(esfera)

                        element = wait.until(EC.visibility_of_element_located((By.XPATH,
                            '//*[@id="content"]/div[2]/section/div/div[3]/div/div[2]/div/form/div[2]/div/div/button[1]')))
                        self.driver.execute_script("arguments[0].click();", element)

                        element = wait.until(EC.visibility_of_element_located((By.XPATH,
                            '//*[@id="content"]/div[2]/section/div/div[3]/div/div[3]/div[2]/div[1]/div/div/button[1]')))
                        self.driver.execute_script("arguments[0].click();", element)

                        time.sleep(4)

                        try:

                            df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                            os.unlink(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                            df_one = df_one[(df_one['TIPO CONTA'] == 'CUSTEIOSUS') | (df_one['TIPO CONTA'] == 'INVESTSUS')]

                            df_estado = pd.concat([df_estado, df_one])

                        except:

                            try:

                                time.sleep(12)

                                df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                                os.unlink(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                                df_one = df_one[(df_one['TIPO CONTA'] == 'CUSTEIOSUS') | (df_one['TIPO CONTA'] == 'INVESTSUS')]

                                df_estado = pd.concat([df_estado, df_one])

                            except:

                                print('Problema na leitura do arquivo!')

                else:

                    element = wait.until(EC.visibility_of_element_located((By.XPATH,
                        '//*[@id="content"]/div[2]/section/div/div[3]/div/div[2]/div/form/div[2]/div/div/button[1]')))
                    self.driver.execute_script("arguments[0].click();", element)

                    element = wait.until(EC.visibility_of_element_located((By.XPATH,
                        '//*[@id="content"]/div[2]/section/div/div[3]/div/div[3]/div[2]/div[1]/div/div/button[1]')))
                    self.driver.execute_script("arguments[0].click();", element)

                    time.sleep(4)

                    try:

                        df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                        os.unlink(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                        df_one = df_one[(df_one['TIPO CONTA'] == 'CUSTEIOSUS') | (df_one['TIPO CONTA'] == 'INVESTSUS')]

                        df_one.insert(loc=2, column='ESFERA', value='MUNICIPAL')

                        df_estado = pd.concat([df_estado, df_one])

                    except:

                        try:

                            time.sleep(12)

                            df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                            os.unlink(path.join(diretorio_dados, 'fns', 'planilha-contas-bancarias.xlsx'))

                            df_one = df_one[(df_one['TIPO CONTA'] == 'CUSTEIOSUS') | (df_one['TIPO CONTA'] == 'INVESTSUS')]

                            df_one.insert(loc=2, column='ESFERA', value='MUNICIPAL')

                            df_estado = pd.concat([df_estado, df_one])

                        except:

                            print('Problema na leitura do arquivo!')

            workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

            writer = pd.ExcelWriter(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'), engine='openpyxl')

            writer.book = workbook

            df_estado.to_excel(writer, sheet_name=estado, index=False)

            writer.save()

            writer.close()

        df_brasil = pd.DataFrame(columns=['UF', 'MUNICÍPIO', 'ESFERA', 'BANCO', 'AGÊNCIA', 'CONTA',
                                          'TIPO CONTA', 'CNPJ', 'ENTIDADE', 'VALOR SALDO'])

        for estado in list(estados_values_capitais.keys()):

            df_estado = pd.read_excel(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'), sheet_name=estado)

            df_brasil = pd.concat([df_brasil, df_estado])

        workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

        writer = pd.ExcelWriter(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'), engine='openpyxl')

        writer.book = workbook

        df_brasil.to_excel(writer, sheet_name='BRASIL', index=False)

        writer.save()

        writer.close()

        workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))

        sheets_names = workbook.sheetnames

        sheets_not_wanted = list(set(sheets_names) - set(['BRASIL']))

        if sheets_not_wanted:

            for sheet in sheets_not_wanted:

                workbook.remove(workbook.get_sheet_by_name(sheet))

        workbook.save(path.join(diretorio_dados, 'fns', 'scraping_FNS_cb.xlsx'))



if __name__ == '__main__':

    print('FNS Contas Bancárias...')
    start_time = time.time()
    scrap_fns_cb = FNS_CONTA_BANCARIA()
    scrap_fns_cb.download()
    print("--- %s segundos ---" % (time.time() - start_time))
