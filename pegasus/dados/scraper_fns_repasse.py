import os
from os import path
import pathlib
from abc import ABC, abstractmethod
import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import xlsxwriter
import openpyxl



# Class implementada pela Monique no repositório CoviDATA
class SeleniumDownloader(ABC):
    def __init__(self, diretorio_dados, url):
        self.driver = self.__configurar_chrome(diretorio_dados)
        self.driver.get(url)

    def download(self):
        self._executar()

        # Aguarda o download
        time.sleep(5)

        #self.driver.close()
        #self.driver.quit()

    @abstractmethod
    def _executar(self):
        pass

    def __configurar_chrome(self, diretorio_dados):
        if not path.exists(diretorio_dados):
            os.makedirs(diretorio_dados)

        chromeOptions = webdriver.ChromeOptions()
        prefs = {"download.default_directory": diretorio_dados}
        chromeOptions.add_experimental_option("prefs", prefs)
        #chromeOptions.add_argument('--headless')
        chromeOptions.add_argument('--start-maximized')


        driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options=chromeOptions)

        driver.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        params = {'cmd': 'Page.setDownloadBehavior',
                  'params': {'behavior': 'allow', 'downloadPath': diretorio_dados}}
        command_result = driver.execute("send_command", params)

        return driver


################################################################################

# Scraper FNS Repasses

################################################################################

diretorio_raiz = pathlib.Path().absolute()
diretorio_dados = diretorio_raiz.joinpath('dados')

estados_values_capitais = {'ACRE': '15',
                           'ALAGOAS': '46',
                           'AMAPA': '6',
                           'AMAZONAS': '37',
                           'BAHIA': '335',
                           'CEARA': '58',
                           'DISTRITO FEDERAL': '0',
                           'ESPIRITO SANTO': '77',
                           'GOIAS': '94',
                           'MARANHAO': '185',
                           'MATO GROSSO': '37',
                           'MATO GROSSO DO SUL': '19',
                           'MINAS GERAIS': '65',
                           'PARA': '18',
                           'PARAIBA': '93',
                           'PARANA': '94',
                           'PERNAMBUCO': '131',
                           'PIAUI': '215',
                           'RIO DE JANEIRO': '67',
                           'RIO GRANDE DO NORTE': '87',
                           'RIO GRANDE DO SUL': '325',
                           'RONDONIA': '36',
                           'RORAIMA': '2',
                           'SANTA CATARINA': '89',
                           'SAO PAULO': '564',
                           'SERGIPE': '2',
                           'TOCANTINS': '88'}

qtd_munic_estado = {'ACRE': 22,
                    'ALAGOAS': 102,
                    'AMAPA': 16,
                    'AMAZONAS': 62,
                    'BAHIA': 417,
                    'CEARA': 184,
                    'DISTRITO FEDERAL': 1,
                    'ESPIRITO SANTO': 78,
                    'GOIAS': 246,
                    'MARANHAO': 217,
                    'MATO GROSSO': 141,
                    'MATO GROSSO DO SUL': 79,
                    'MINAS GERAIS': 853,
                    'PARA': 144,
                    'PARAIBA': 223,
                    'PARANA': 399,
                    'PERNAMBUCO': 184,
                    'PIAUI': 224,
                    'RIO DE JANEIRO': 92,
                    'RIO GRANDE DO NORTE': 167,
                    'RIO GRANDE DO SUL': 497,
                    'RONDONIA': 52,
                    'RORAIMA': 15,
                    'SANTA CATARINA': 295,
                    'SAO PAULO': 645,
                    'SERGIPE': 75,
                    'TOCANTINS': 139}

# Define a classe referida como herdeira da class "SeleniumDownloader"
class FNS_REPASSE(SeleniumDownloader):

    # Sobrescreve o construtor da class "SeleniumDownloader"
    def __init__(self):
        super().__init__(path.join(diretorio_dados, 'fns'),
                         'https://consultafns.saude.gov.br/#/consolidada')

    # Implementa localmente o método interno e vazio da class "SeleniumDownloader"
    def _executar(self):

        wait = WebDriverWait(self.driver, 20)

        time.sleep(5)

        for tipo in ['covid19', 'geral']:

            try:

                workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))

                workbook.save(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))

            except:

                workbook = xlsxwriter.Workbook(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))

                workbook.close()

            workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))

            sheets_names = workbook.sheetnames

            number_sheets = len(sheets_names)

            workbook.save(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))


        select = Select(self.driver.find_element_by_id('ano'))
        select.select_by_visible_text('2020')

        for estado in list(estados_values_capitais.keys())[number_sheets - 1:]:

            print(estado)

            df_estado_covid = pd.DataFrame(columns=['UF', 'MUNICIPIO', 'ENTIDADE',
                                                    'BLOCO', 'GRUPO', 'VALOR LIQUIDO'])

            df_estado_geral = pd.DataFrame(columns=['UF', 'MUNICIPIO', 'ENTIDADE',
                                                    'BLOCO', 'VALOR LIQUIDO'])

            time.sleep(5)

            select = Select(self.driver.find_element_by_id('estado'))
            select.select_by_visible_text(estado)

            qtd_estado = qtd_munic_estado[estado]

            for municipio in range(qtd_estado):

                time.sleep(1)

                municipio = str(municipio)

                select = Select(self.driver.find_element_by_id('municipio'))
                select.select_by_value(municipio)

                if municipio == estados_values_capitais[estado]:

                    for esfera in ['Estadual', 'Municipal']:

                        time.sleep(1)

                        select = Select(self.driver.find_element_by_id('tipoRepasse'))
                        select.select_by_visible_text(esfera)

                        element = wait.until(EC.visibility_of_element_located((By.XPATH,
                            '//*[@id="repasses"]/div/form/div[2]/div/div/button[1]')))
                        self.driver.execute_script("arguments[0].click();", element)

                        element = wait.until(EC.visibility_of_element_located((By.XPATH,
                            '//*[@id="content"]/div[2]/section/div/div[3]/div/div[3]/div[2]/div[4]/div/div/button[1]')))
                        self.driver.execute_script("arguments[0].click();", element)

                        time.sleep(5)

                        try:

                            df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                            os.unlink(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                            df_one.rename(columns={'Valor Líquido': 'VALOR LIQUIDO'}, inplace=True)

                            df_one['BLOCO'].replace(
                                'Manutencao das Acoes e Servicos Publicos de Saude (CUSTEIO)', 'CUSTEIO', inplace=True)
                            df_one['BLOCO'].replace(
                                'Estruturacao da Rede de Servicos Publicos de Saude (INVESTIMENTO)', 'INVESTIMENTO', inplace=True)

                            df_covid = df_one[df_one['GRUPO'].str.contains('^CORONAVIRUS', regex=True)]

                            result_groupby_covid = \
                                df_covid.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO', 'GRUPO'], as_index=False)['VALOR LIQUIDO'].sum()

                            df_estado_covid = pd.concat([df_estado_covid, result_groupby_covid])

                            df_geral = df_one.drop(columns=['GRUPO'])

                            result_groupby_geral = \
                                df_geral.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO'], as_index=False)['VALOR LIQUIDO'].sum()

                            df_estado_geral = pd.concat([df_estado_geral, result_groupby_geral])

                        except:

                            try:

                                print('Tentando novamente...')

                                time.sleep(20)

                                df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                                os.unlink(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                                df_one.rename(columns={'Valor Líquido': 'VALOR LIQUIDO'}, inplace=True)

                                df_one['BLOCO'].replace(
                                    'Manutencao das Acoes e Servicos Publicos de Saude (CUSTEIO)', 'CUSTEIO', inplace=True)
                                df_one['BLOCO'].replace(
                                    'Estruturacao da Rede de Servicos Publicos de Saude (INVESTIMENTO)', 'INVESTIMENTO', inplace=True)

                                df_covid = df_one[df_one['GRUPO'].str.contains('^CORONAVIRUS', regex=True)]

                                result_groupby_covid = \
                                    df_covid.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO', 'GRUPO'], as_index=False)['VALOR LIQUIDO'].sum()

                                df_estado_covid = pd.concat([df_estado_covid, result_groupby_covid])

                                df_geral = df_one.drop(columns=['GRUPO'])

                                result_groupby_geral = \
                                    df_geral.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO'], as_index=False)['VALOR LIQUIDO'].sum()

                                df_estado_geral = pd.concat([df_estado_geral, result_groupby_geral])

                            except:

                                print('Problema na leitura do arquivo!')

                else:

                    element = wait.until(EC.visibility_of_element_located((By.XPATH,
                        '//*[@id="repasses"]/div/form/div[2]/div/div/button[1]')))
                    self.driver.execute_script("arguments[0].click();", element)

                    element = wait.until(EC.visibility_of_element_located((By.XPATH,
                        '//*[@id="content"]/div[2]/section/div/div[3]/div/div[3]/div[2]/div[4]/div/div/button[1]')))
                    self.driver.execute_script("arguments[0].click();", element)

                    time.sleep(5)

                    try:

                        df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                        os.unlink(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                        df_one.rename(columns={'Valor Líquido': 'VALOR LIQUIDO'}, inplace=True)

                        df_one['BLOCO'].replace(
                            'Manutencao das Acoes e Servicos Publicos de Saude (CUSTEIO)', 'CUSTEIO', inplace=True)
                        df_one['BLOCO'].replace(
                            'Estruturacao da Rede de Servicos Publicos de Saude (INVESTIMENTO)', 'INVESTIMENTO', inplace=True)

                        df_covid = df_one[df_one['GRUPO'].str.contains('^CORONAVIRUS', regex=True)]

                        result_groupby_covid = \
                            df_covid.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO', 'GRUPO'], as_index=False)['VALOR LIQUIDO'].sum()

                        df_estado_covid = pd.concat([df_estado_covid, result_groupby_covid])

                        df_geral = df_one.drop(columns=['GRUPO'])

                        result_groupby_geral = \
                            df_geral.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO'], as_index=False)['VALOR LIQUIDO'].sum()

                        df_estado_geral = pd.concat([df_estado_geral, result_groupby_geral])

                    except:

                        try:

                            print('Tentando novamente...')

                            time.sleep(20)

                            df_one = pd.read_excel(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                            os.unlink(path.join(diretorio_dados, 'fns', 'consulta-consolidada-planilha.xlsx'))

                            df_one.rename(columns={'Valor Líquido': 'VALOR LIQUIDO'}, inplace=True)

                            df_one['BLOCO'].replace(
                                'Manutencao das Acoes e Servicos Publicos de Saude (CUSTEIO)', 'CUSTEIO', inplace=True)
                            df_one['BLOCO'].replace(
                                'Estruturacao da Rede de Servicos Publicos de Saude (INVESTIMENTO)', 'INVESTIMENTO', inplace=True)

                            df_covid = df_one[df_one['GRUPO'].str.contains('^CORONAVIRUS', regex=True)]

                            result_groupby_covid = \
                                df_covid.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO', 'GRUPO'], as_index=False)['VALOR LIQUIDO'].sum()

                            df_estado_covid = pd.concat([df_estado_covid, result_groupby_covid])

                            df_geral = df_one.drop(columns=['GRUPO'])

                            result_groupby_geral = \
                                df_geral.groupby(['UF', 'MUNICIPIO', 'ENTIDADE', 'BLOCO'], as_index=False)['VALOR LIQUIDO'].sum()

                            df_estado_geral = pd.concat([df_estado_geral, result_groupby_geral])

                        except:

                            print('Problema na leitura do arquivo!')

            for tipo in ['covid19', 'geral']:

                workbook = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'))

                writer = pd.ExcelWriter(path.join(diretorio_dados, 'fns', f'scraping_FNS_repasse_{tipo}.xlsx'), engine='openpyxl')

                writer.book = workbook

                if tipo == 'covid19':

                    df_estado_covid.to_excel(writer, sheet_name=estado, index=False)

                else:

                    df_estado_geral.to_excel(writer, sheet_name=estado, index=False)

                writer.save()

                writer.close()

        df_brasil_covid = pd.DataFrame(columns=['UF', 'MUNICIPIO', 'ENTIDADE',
                                                'BLOCO', 'GRUPO', 'VALOR LIQUIDO'])

        df_brasil_geral = pd.DataFrame(columns=['UF', 'MUNICIPIO', 'ENTIDADE',
                                                'BLOCO', 'VALOR LIQUIDO'])

        for estado in list(estados_values_capitais.keys()):

            df_estado_covid = pd.read_excel(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_covid19.xlsx'), sheet_name=estado)

            df_estado_geral = pd.read_excel(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_geral.xlsx'), sheet_name=estado)

            df_brasil_covid = pd.concat([df_brasil_covid, df_estado_covid])

            df_brasil_geral = pd.concat([df_brasil_geral, df_estado_geral])

        workbook1 = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_covid19.xlsx'))

        writer1 = pd.ExcelWriter(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_covid19.xlsx'), engine='openpyxl')

        writer1.book = workbook1

        df_brasil_covid.to_excel(writer1, sheet_name='BRASIL-COVID19', index=False)

        workbook2 = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_geral.xlsx'))

        writer1.save()

        writer1.close()

        writer2 = pd.ExcelWriter(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_geral.xlsx'), engine='openpyxl')

        writer2.book = workbook2

        df_brasil_geral.to_excel(writer2, sheet_name='BRASIL-GERAL', index=False)

        writer2.save()

        writer2.close()

        workbook1 = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_covid19.xlsx'))

        sheets_names1 = workbook1.sheetnames

        sheets_not_wanted1 = list(set(sheets_names1) - set(['BRASIL-COVID19']))

        if sheets_not_wanted1:

            for sheet in sheets_not_wanted1:

                workbook1.remove(workbook1.get_sheet_by_name(sheet))

        workbook1.save(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_covid19.xlsx'))

        workbook2 = openpyxl.load_workbook(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_geral.xlsx'))

        sheets_names2 = workbook2.sheetnames

        sheets_not_wanted2 = list(set(sheets_names2) - set(['BRASIL-GERAL']))

        if sheets_not_wanted2:

            for sheet in sheets_not_wanted2:

                workbook2.remove(workbook2.get_sheet_by_name(sheet))

        workbook2.save(path.join(diretorio_dados, 'fns', 'scraping_FNS_repasse_geral.xlsx'))



if __name__ == '__main__':

    print('FNS Repasses...')
    start_time = time.time()
    scrap_fns_repasse = FNS_REPASSE()
    scrap_fns_repasse.download()
    print("--- %s segundos ---" % (time.time() - start_time))
